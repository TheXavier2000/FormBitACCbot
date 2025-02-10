import pandas as pd
import openpyxl
import unicodedata
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def limpiar_nombre(nombre):
    """Elimina tildes y normaliza caracteres especiales en los nombres."""
    return ''.join(c for c in unicodedata.normalize('NFKD', nombre) if not unicodedata.combining(c))

def generar_csv_con_formato(df):
    """Genera un archivo Excel con formato mejorado a partir de un DataFrame"""
    
    # Guardar como archivo Excel
    ruta_salida = "reporte_formateado.xlsx"
    df.to_excel(ruta_salida, index=False)

    # Cargar el archivo Excel con openpyxl
    wb = openpyxl.load_workbook(ruta_salida)
    ws = wb.active

    # Aplicar estilo a la primera fila (encabezados)
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")  # Azul oscuro
    header_font = Font(color="FFFFFF", bold=True)  # Texto blanco y negrita
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_style
    
    # Aplicar colores alternos a las filas de datos
    fill_azul_claro = PatternFill(start_color="CCECFF", end_color="CCECFF", fill_type="solid")
    fill_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
        fill = fill_azul_claro if i % 2 == 0 else fill_blanco
        for cell in row:
            cell.fill = fill
            cell.border = border_style
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar con un margen extra
    
    # Guardar cambios
    wb.save(ruta_salida)
    print(f"✅ Archivo generado correctamente: {ruta_salida}")
